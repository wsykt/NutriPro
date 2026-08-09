#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中国食物成分数据处理脚本 v2
从开源项目提取关键字段，根据分类控制字段可见性，生成审核用Excel文件
"""

import os
import json
import re
from datetime import datetime

PROJECT_ROOT = r"C:\Users\13425\Desktop\个人健康助手"
DATA_DIR = os.path.join(PROJECT_ROOT, "china-food-data", "json_data_vision_251206_Qwen2-5-VL-72B-Instruct")
GI_FILE = os.path.join(PROJECT_ROOT, "china-food-data", "json_gi_of_foods", "glycemic_index_of_foods.json")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "health", "ai_service", "test_output", "food_data_review")
os.makedirs(OUTPUT_DIR, exist_ok=True)

FILE_CATEGORY_MAP = {
    "谷类及其制品": "主食",
    "蔬菜类及其制品": "蔬菜",
    "水果类及其制品": "水果",
    "畜肉类及其制品": "肉蛋类",
    "禽肉类及其制品": "肉蛋类",
    "鱼虾蟹贝类": "水产",
    "蛋类及其制品": "肉蛋类",
    "乳类及其制品": "奶类",
    "干豆类及其制品": "豆制品",
    "坚果种子类": "零食",
    "薯类淀粉及其制品": "主食",
    "菌藻类": "蔬菜",
    "植物油": "油脂类",
    "动物油脂类": "油脂类",
    "其他类": "零食",
}

FIELD_VISIBILITY = {
    "主食":       {"show_gi": True,  "show_folic_acid": True,  "show_dha": False},
    "肉蛋类":     {"show_gi": False, "show_folic_acid": False, "show_dha": False},
    "水产":       {"show_gi": False, "show_folic_acid": True,  "show_dha": True},
    "蔬菜":       {"show_gi": True,  "show_folic_acid": True,  "show_dha": False},
    "水果":       {"show_gi": True,  "show_folic_acid": False, "show_dha": False},
    "豆制品":     {"show_gi": True,  "show_folic_acid": True,  "show_dha": False},
    "奶类":       {"show_gi": True,  "show_folic_acid": False, "show_dha": False},
    "油脂类":     {"show_gi": False, "show_folic_acid": False, "show_dha": False},
    "零食":       {"show_gi": True,  "show_folic_acid": False, "show_dha": False},
    "其他":       {"show_gi": False, "show_folic_acid": False, "show_dha": False},
}

FIELD_VISIBILITY_DESC = {
    "主食":       "GI✅ 叶酸✅ DHA❌",
    "肉蛋类":     "GI❌ 叶酸❌ DHA❌",
    "水产":       "GI❌ 叶酸✅ DHA✅",
    "蔬菜":       "GI✅ 叶酸✅ DHA❌",
    "水果":       "GI✅ 叶酸❌ DHA❌",
    "豆制品":     "GI✅ 叶酸✅ DHA❌",
    "奶类":       "GI✅ 叶酸❌ DHA❌",
    "油脂类":     "GI❌ 叶酸❌ DHA❌",
    "零食":       "GI✅ 叶酸❌ DHA❌",
    "其他":       "GI❌ 叶酸❌ DHA❌",
}


def parse_float(value):
    if not value:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in ['-', '', 'Tr', 'tr', 'N/A', 'NA', '痕量']:
        return None
    s = re.sub(r'[^\d.]', '', s)
    if not s or s == '.':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_gi_data():
    gi_data = {}
    if not os.path.exists(GI_FILE):
        print(f"GI文件不存在: {GI_FILE}")
        return gi_data

    with open(GI_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    for group in data:
        for item in group.get('list', []):
            food_name = item.get('foodName', '').strip()
            food_name = food_name.replace('*', '').strip()
            gi = parse_float(item.get('GI'))
            if food_name and gi:
                gi_data[food_name] = gi
                name_parts = food_name.split('（')
                if len(name_parts) > 1:
                    base_name = name_parts[0].strip()
                    if base_name not in gi_data:
                        gi_data[base_name] = gi
                name_parts2 = food_name.split('(')
                if len(name_parts2) > 1:
                    base_name = name_parts2[0].strip()
                    if base_name not in gi_data:
                        gi_data[base_name] = gi

    print(f"加载GI数据: {len(gi_data)} 条")
    return gi_data


def get_category_from_filename(filename):
    for key, value in FILE_CATEGORY_MAP.items():
        if key in filename:
            return value
    return "其他"


def find_gi_value(food_name, gi_data):
    if food_name in gi_data:
        return gi_data[food_name]

    clean_name = food_name.replace('*', '').strip()
    if clean_name in gi_data:
        return gi_data[clean_name]

    for sep in ['（', '(']:
        name_parts = clean_name.split(sep)
        if len(name_parts) > 1:
            base_name = name_parts[0].strip()
            if base_name in gi_data:
                return gi_data[base_name]

    return None


def get_field_visibility(category):
    return FIELD_VISIBILITY.get(category, FIELD_VISIBILITY["其他"])


def extract_data():
    print("=" * 60)
    print("中国食物成分数据处理 v2")
    print("=" * 60)

    gi_data = load_gi_data()

    json_files = sorted([f for f in os.listdir(DATA_DIR) if f.endswith('.json')])
    print(f"\n发现 {len(json_files)} 个数据文件")

    all_foods = []
    stats = {
        "total": 0,
        "categories": {},
        "total_calorie_zero": 0,
        "show_gi_count": 0,
        "show_folic_acid_count": 0,
        "show_dha_count": 0,
        "with_gi_and_show": 0,
        "field_visibility_by_cat": {},
        "missing_gi_samples": [],
    }

    for json_file in json_files:
        filepath = os.path.join(DATA_DIR, json_file)
        category = get_category_from_filename(json_file)
        visibility = get_field_visibility(category)

        if category not in stats["field_visibility_by_cat"]:
            stats["field_visibility_by_cat"][category] = {
                "show_gi": 0, "show_folic_acid": 0, "show_dha": 0,
                "with_gi_value": 0, "missing_gi": 0,
            }

        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        for item in data:
            food_name = item.get('foodName', '').strip()
            if not food_name:
                continue

            calorie = parse_float(item.get('energyKCal'))
            protein = parse_float(item.get('protein'))
            fat = parse_float(item.get('fat'))
            carb = parse_float(item.get('CHO'))
            diet_fiber = parse_float(item.get('dietaryFiber'))
            calcium = parse_float(item.get('Ca'))

            gi_value = find_gi_value(food_name, gi_data)

            show_gi = visibility["show_gi"]
            show_folic_acid = visibility["show_folic_acid"]
            show_dha = visibility["show_dha"]

            food_record = {
                "food_name": food_name,
                "food_category": category,
                "calorie": calorie,
                "protein": protein,
                "fat": fat,
                "carb": carb,
                "diet_fiber": diet_fiber,
                "calcium": calcium,
                "gi_value": gi_value if show_gi else None,
                "folic_acid": None,
                "dha": None,
                "show_gi": show_gi,
                "show_folic_acid": show_folic_acid,
                "show_dha": show_dha,
            }

            all_foods.append(food_record)
            stats["total"] += 1

            if category not in stats["categories"]:
                stats["categories"][category] = 0
            stats["categories"][category] += 1

            if show_gi:
                stats["show_gi_count"] += 1
                stats["field_visibility_by_cat"][category]["show_gi"] += 1
                if gi_value is not None:
                    stats["with_gi_and_show"] += 1
                    stats["field_visibility_by_cat"][category]["with_gi_value"] += 1
                else:
                    stats["field_visibility_by_cat"][category]["missing_gi"] += 1
                    if len(stats["missing_gi_samples"]) < 10:
                        stats["missing_gi_samples"].append(food_name)

            if show_folic_acid:
                stats["show_folic_acid_count"] += 1
                stats["field_visibility_by_cat"][category]["show_folic_acid"] += 1

            if show_dha:
                stats["show_dha_count"] += 1
                stats["field_visibility_by_cat"][category]["show_dha"] += 1

            if calorie is None or calorie == 0:
                stats["total_calorie_zero"] += 1

    print(f"\n提取完成:")
    print(f"  总记录数: {stats['total']}")
    print(f"  ├─ 需要显示GI值: {stats['show_gi_count']} ({stats['show_gi_count']/max(stats['total'],1)*100:.1f}%)")
    print(f"  │   └─ 有GI值: {stats['with_gi_and_show']}")
    print(f"  │   └─ 无GI值（待补充）: {stats['show_gi_count'] - stats['with_gi_and_show']}")
    print(f"  ├─ 需要显示叶酸: {stats['show_folic_acid_count']} ({stats['show_folic_acid_count']/max(stats['total'],1)*100:.1f}%)")
    print(f"  └─ 需要显示DHA: {stats['show_dha_count']} ({stats['show_dha_count']/max(stats['total'],1)*100:.1f}%)")
    print(f"  热量为0或缺失: {stats['total_calorie_zero']}")

    print(f"\n分类分布与字段可见性:")
    for cat, count in sorted(stats["categories"].items(), key=lambda x: -x[1]):
        vis = FIELD_VISIBILITY_DESC.get(cat, "")
        s = stats["field_visibility_by_cat"].get(cat, {})
        gi_info = f", GI有值={s.get('with_gi_value',0)}" if s.get("show_gi", 0) > 0 else ""
        print(f"  {cat:8s}: {count:4d} | {vis}{gi_info}")

    print(f"\n需显示GI但无数据的示例（前10个）:")
    for name in stats["missing_gi_samples"]:
        print(f"  {name}")

    return all_foods, stats


def generate_csv(all_foods, stats):
    csv_path = os.path.join(OUTPUT_DIR, "china_food_simplified.csv")

    headers = [
        "food_name",
        "food_category",
        "calorie(kcal/100g)",
        "protein(g/100g)",
        "fat(g/100g)",
        "carb(g/100g)",
        "diet_fiber(g/100g)",
        "calcium(mg/100g)",
        "gi_value",
        "folic_acid(μg)",
        "dha(mg)",
        "show_gi",
        "show_folic_acid",
        "show_dha",
    ]

    with open(csv_path, 'w', encoding='utf-8-sig') as f:
        f.write(','.join(headers) + '\n')
        for food in all_foods:
            row = [
                food["food_name"],
                food["food_category"],
                str(food["calorie"] or ''),
                str(food["protein"] or ''),
                str(food["fat"] or ''),
                str(food["carb"] or ''),
                str(food["diet_fiber"] or ''),
                str(food["calcium"] or ''),
                str(food["gi_value"] or ''),
                str(food["folic_acid"] or ''),
                str(food["dha"] or ''),
                str(food["show_gi"]),
                str(food["show_folic_acid"]),
                str(food["show_dha"]),
            ]
            f.write(','.join(row) + '\n')

    print(f"\nCSV文件已生成: {csv_path}")
    return csv_path


def generate_excel_with_summary(all_foods, stats):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 未安装，跳过Excel生成，仅生成CSV")
        return None

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "数据概览"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="left", vertical="center")
    number_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws_summary['A1'] = "中国食物成分数据 v2 - 审核报告"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:F1')

    summary_data = [
        ["指标", "数值", "说明"],
        ["总记录数", stats["total"], "从开源项目提取的食物数量"],
        ["需显示GI值", f"{stats['show_gi_count']} ({stats['show_gi_count']/max(stats['total'],1)*100:.1f}%)", "主食/蔬菜/水果/豆制品/奶类/零食"],
        ["  其中有GI值", stats["with_gi_and_show"], "从GI数据库匹配到的"],
        ["  其中无GI值", stats["show_gi_count"] - stats["with_gi_and_show"], "需人工补充或接受无数据"],
        ["需显示叶酸", f"{stats['show_folic_acid_count']} ({stats['show_folic_acid_count']/max(stats['total'],1)*100:.1f}%)", "主食/蔬菜/水果/豆制品/奶类"],
        ["需显示DHA", f"{stats['show_dha_count']} ({stats['show_dha_count']/max(stats['total'],1)*100:.1f}%)", "仅水产类"],
        ["热量缺失/为0", stats["total_calorie_zero"], "无法提供热量信息的记录"],
        ["分类数", len(stats["categories"]), "覆盖的食物分类数量"],
    ]

    for i, row in enumerate(summary_data, start=3):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 3:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            elif j == 2 and isinstance(value, (int, float)):
                cell.alignment = number_align
            else:
                cell.alignment = cell_align

    ws_summary['A14'] = "分类分布与字段可见性"
    ws_summary['A14'].font = Font(bold=True, size=12)

    cat_header = ["分类", "数量", "占比", "GI可见", "叶酸可见", "DHA可见"]
    for j, value in enumerate(cat_header, start=1):
        cell = ws_summary.cell(row=15, column=j, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_idx = 16
    for cat, count in sorted(stats["categories"].items(), key=lambda x: -x[1]):
        vis = get_field_visibility(cat)
        ws_summary.cell(row=row_idx, column=1, value=cat).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=count).border = thin_border
        ws_summary.cell(row=row_idx, column=2).alignment = number_align
        ws_summary.cell(row=row_idx, column=3, value=f"{count/max(stats['total'],1)*100:.1f}%").border = thin_border
        ws_summary.cell(row=row_idx, column=3).alignment = number_align
        ws_summary.cell(row=row_idx, column=4, value="是" if vis["show_gi"] else "否").border = thin_border
        ws_summary.cell(row=row_idx, column=4).alignment = center_align
        ws_summary.cell(row=row_idx, column=5, value="是" if vis["show_folic_acid"] else "否").border = thin_border
        ws_summary.cell(row=row_idx, column=5).alignment = center_align
        ws_summary.cell(row=row_idx, column=6, value="是" if vis["show_dha"] else "否").border = thin_border
        ws_summary.cell(row=row_idx, column=6).alignment = center_align
        row_idx += 1

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 35
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 10

    ws_data = wb.create_sheet("食物数据")

    data_headers = [
        "序号",
        "食物名称",
        "分类",
        "热量(kcal)",
        "蛋白质(g)",
        "脂肪(g)",
        "碳水(g)",
        "膳食纤维(g)",
        "钙(mg)",
        "GI值",
        "叶酸(μg)",
        "DHA(mg)",
        "显示GI",
        "显示叶酸",
        "显示DHA",
        "需要审核",
        "备注",
    ]

    for j, value in enumerate(data_headers, start=1):
        cell = ws_data.cell(row=1, column=j, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, food in enumerate(all_foods, start=2):
        ws_data.cell(row=i, column=1, value=i-1).border = thin_border
        ws_data.cell(row=i, column=2, value=food["food_name"]).border = thin_border
        ws_data.cell(row=i, column=3, value=food["food_category"]).border = thin_border

        ws_data.cell(row=i, column=4, value=food["calorie"]).border = thin_border
        ws_data.cell(row=i, column=4).alignment = number_align
        ws_data.cell(row=i, column=5, value=food["protein"]).border = thin_border
        ws_data.cell(row=i, column=5).alignment = number_align
        ws_data.cell(row=i, column=6, value=food["fat"]).border = thin_border
        ws_data.cell(row=i, column=6).alignment = number_align
        ws_data.cell(row=i, column=7, value=food["carb"]).border = thin_border
        ws_data.cell(row=i, column=7).alignment = number_align
        ws_data.cell(row=i, column=8, value=food["diet_fiber"]).border = thin_border
        ws_data.cell(row=i, column=8).alignment = number_align
        ws_data.cell(row=i, column=9, value=food["calcium"]).border = thin_border
        ws_data.cell(row=i, column=9).alignment = number_align

        if food["show_gi"]:
            ws_data.cell(row=i, column=10, value=food["gi_value"]).border = thin_border
            ws_data.cell(row=i, column=10).alignment = number_align
        else:
            ws_data.cell(row=i, column=10, value="-").border = thin_border
            ws_data.cell(row=i, column=10).alignment = center_align

        if food["show_folic_acid"]:
            ws_data.cell(row=i, column=11, value=food["folic_acid"] if food["folic_acid"] is not None else "-").border = thin_border
            ws_data.cell(row=i, column=11).alignment = number_align if food["folic_acid"] is not None else center_align
        else:
            ws_data.cell(row=i, column=11, value="-").border = thin_border
            ws_data.cell(row=i, column=11).alignment = center_align

        if food["show_dha"]:
            ws_data.cell(row=i, column=12, value=food["dha"] if food["dha"] is not None else "-").border = thin_border
            ws_data.cell(row=i, column=12).alignment = number_align if food["dha"] is not None else center_align
        else:
            ws_data.cell(row=i, column=12, value="-").border = thin_border
            ws_data.cell(row=i, column=12).alignment = center_align

        ws_data.cell(row=i, column=13, value="是" if food["show_gi"] else "否").border = thin_border
        ws_data.cell(row=i, column=13).alignment = center_align
        ws_data.cell(row=i, column=14, value="是" if food["show_folic_acid"] else "否").border = thin_border
        ws_data.cell(row=i, column=14).alignment = center_align
        ws_data.cell(row=i, column=15, value="是" if food["show_dha"] else "否").border = thin_border
        ws_data.cell(row=i, column=15).alignment = center_align

        need_review = ""
        notes = []
        if food["show_gi"] and food["gi_value"] is None:
            need_review = "是"
            notes.append("需显示GI但无数据")
        if food["calorie"] is None or food["calorie"] == 0:
            need_review = "是"
            notes.append("热量缺失")
        if food["protein"] is None and food["fat"] is None and food["carb"] is None:
            need_review = "是"
            notes.append("三大营养素全缺")

        ws_data.cell(row=i, column=16, value=need_review).border = thin_border
        ws_data.cell(row=i, column=17, value="; ".join(notes)).border = thin_border

    col_widths = [8, 25, 10, 12, 12, 12, 12, 14, 10, 10, 10, 10, 8, 8, 8, 12, 25]
    for idx, width in enumerate(col_widths, start=1):
        ws_data.column_dimensions[get_column_letter(idx)].width = width

    excel_path = os.path.join(OUTPUT_DIR, "china_food_simplified_review.xlsx")
    wb.save(excel_path)
    print(f"\nExcel文件已生成: {excel_path}")
    return excel_path


def generate_review_report(all_foods, stats):
    report_path = os.path.join(OUTPUT_DIR, "数据处理报告_v2.md")

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("# 中国食物成分数据 v2 - 审核报告\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"**数据来源**: https://github.com/Sanotsu/china-food-composition-data\n\n")
        f.write(f"**原始数据版本**: json_data_vision_251206_Qwen2-5-VL-72B-Instruct\n\n")

        f.write("## 一、处理概述\n\n")
        f.write("本次处理从开源项目《中国食物成分表标准版（第6版）》数据中提取关键字段，**根据食物分类控制字段可见性**，简化后供审核。\n\n")

        f.write("### 1.1 核心设计理念\n\n")
        f.write("并非所有营养字段对所有食材都有意义。例如：\n\n")
        f.write("- **GI值**：对主食、蔬菜、水果、豆制品、奶类、零食有意义；对肉类、水产、油脂类无意义\n")
        f.write("- **叶酸**：对主食、蔬菜、豆制品、水产有意义（尤其深绿蔬菜与动物肝脏/海鱼）；对水果、奶类、肉类无意义\n")
        f.write("- **DHA**：仅对水产类（海鱼）有意义；其他食材几乎不含DHA\n\n")

        f.write("### 1.2 字段可见性矩阵\n\n")
        f.write("| 分类 | GI值 | 叶酸 | DHA | 说明 |\n")
        f.write("|------|------|------|-----|------|\n")
        for cat in ["主食", "肉蛋类", "水产", "蔬菜", "水果", "豆制品", "奶类", "油脂类", "零食", "其他"]:
            vis = FIELD_VISIBILITY[cat]
            gi_vis = "✅" if vis["show_gi"] else "❌"
            fa_vis = "✅" if vis["show_folic_acid"] else "❌"
            dha_vis = "✅" if vis["show_dha"] else "❌"
            f.write(f"| {cat} | {gi_vis} | {fa_vis} | {dha_vis} | {FIELD_VISIBILITY_DESC.get(cat, '')} |\n")

        f.write("\n### 1.3 提取字段说明\n\n")
        f.write("| 目标字段 | 开源字段 | 单位 | 是否存在于开源数据 |\n")
        f.write("|----------|----------|------|------------------|\n")
        f.write("| food_name | foodName | - | ✅ 有 |\n")
        f.write("| food_category | 文件名映射 | - | ✅ 有 |\n")
        f.write("| calorie | energyKCal | kcal/100g | ✅ 有 |\n")
        f.write("| protein | protein | g/100g | ✅ 有 |\n")
        f.write("| fat | fat | g/100g | ✅ 有 |\n")
        f.write("| carb | CHO | g/100g | ✅ 有 |\n")
        f.write("| diet_fiber | dietaryFiber | g/100g | ✅ 有 |\n")
        f.write("| calcium | Ca | mg/100g | ✅ 有 |\n")
        f.write("| gi_value | 外部GI库 | - | ⚠️ 需匹配 |\n")
        f.write("| folic_acid | - | μg/100g | ❌ 无此字段 |\n")
        f.write("| dha | - | mg/100g | ❌ 无此字段 |\n\n")

        f.write("### 1.4 show_gi / show_folic_acid / show_dha 标记说明\n\n")
        f.write("这三个布尔字段用于控制前端是否显示对应的营养指标：\n\n")
        f.write("- **show_gi = True**：前端应显示GI值列\n")
        f.write("- **show_gi = False**：前端应隐藏GI值列（显示 '-'）\n")
        f.write("- **show_folic_acid = True**：前端应显示叶酸列\n")
        f.write("- **show_folic_acid = False**：前端应隐藏叶酸列\n")
        f.write("- **show_dha = True**：前端应显示DHA列\n")
        f.write("- **show_dha = False**：前端应隐藏DHA列\n\n")

        f.write("### 1.5 未提取字段（已移除）\n\n")
        f.write("以下字段因非核心营养指标，已从处理结果中移除：\n\n")
        f.write("- 水分 (water), 可食部分 (edible), 能量千焦 (energyKJ)\n")
        f.write("- 胆固醇 (cholesterol), 灰分 (ash)\n")
        f.write("- 维生素A (vitaminA), 胡萝卜素 (carotene), 视黄醇 (retinol)\n")
        f.write("- 硫胺素 (thiamin), 核黄素 (riboflavin), 尼克酸 (niacin)\n")
        f.write("- 维生素C (vitaminC), 维生素E (vitaminETotal/E1/E2/E3)\n")
        f.write("- 磷 (P), 钾 (K), 钠 (Na), 镁 (Mg), 铁 (Fe), 锌 (Zn)\n")
        f.write("- 硒 (Se), 铜 (Cu), 锰 (Mn)\n\n")

        f.write("## 二、数据统计\n\n")
        f.write("### 2.1 总体情况\n\n")
        f.write("| 指标 | 数值 |\n")
        f.write("|------|------|\n")
        f.write(f"| 总记录数 | {stats['total']} |\n")
        f.write(f"| 需显示GI值 | {stats['show_gi_count']} ({stats['show_gi_count']/max(stats['total'],1)*100:.1f}%) |\n")
        f.write(f"|   其中有GI值 | {stats['with_gi_and_show']} |\n")
        f.write(f"|   其中无GI值（待补充） | {stats['show_gi_count'] - stats['with_gi_and_show']} |\n")
        f.write(f"| 需显示叶酸 | {stats['show_folic_acid_count']} ({stats['show_folic_acid_count']/max(stats['total'],1)*100:.1f}%) | 主食/蔬菜/豆制品/水产 |\n")
        f.write(f"| 需显示DHA | {stats['show_dha_count']} ({stats['show_dha_count']/max(stats['total'],1)*100:.1f}%) | 仅水产类 |\n")
        f.write(f"| 热量缺失/为0 | {stats['total_calorie_zero']} |\n\n")

        f.write("### 2.2 分类分布与字段可见性\n\n")
        f.write("| 分类 | 数量 | 占比 | GI可见 | 叶酸可见 | DHA可见 |\n")
        f.write("|------|------|------|--------|----------|--------|\n")
        for cat, count in sorted(stats["categories"].items(), key=lambda x: -x[1]):
            vis = get_field_visibility(cat)
            gi_vis = "✅" if vis["show_gi"] else "❌"
            fa_vis = "✅" if vis["show_folic_acid"] else "❌"
            dha_vis = "✅" if vis["show_dha"] else "❌"
            f.write(f"| {cat} | {count} | {count/max(stats['total'],1)*100:.1f}% | {gi_vis} | {fa_vis} | {dha_vis} |\n")

        f.write("\n## 三、审核要点\n\n")
        f.write("### 3.1 需要重点审核的问题\n\n")
        f.write("1. **GI值缺失**：需显示GI值的食材中，部分没有匹配到GI数据，需确认是否接受此现状\n")
        f.write("2. **热量数据**：部分记录热量为0或缺失，需确认数据质量\n")
        f.write("3. **字段可见性矩阵**：请审核GI/叶酸/DHA的可见性规则是否符合预期\n")
        f.write("4. **分类准确性**：基于文件名的自动分类可能不完全准确，建议抽查\n")
        f.write("5. **叶酸/DHA数据**：这两个字段在开源数据中不存在，建议先保留为空，后续从专业数据源补充\n\n")

        f.write("### 3.2 建议的审核流程\n\n")
        f.write("1. 查看Excel文件中的\"数据概览\"sheet，确认总体统计\n")
        f.write("2. 查看\"分类分布与字段可见性\"表，确认可见性规则\n")
        f.write("3. 切换到\"食物数据\"sheet\n")
        f.write("4. 筛选\"需要审核\"列为\"是\"的记录\n")
        f.write("5. 抽查各分类下的代表性食物，确认分类和字段可见性正确\n")
        f.write("6. 确认数据质量无误后，再考虑导入数据库\n\n")

        f.write("## 四、输出文件\n\n")
        f.write(f"- **Excel审核文件**: `china_food_simplified_review.xlsx`\n")
        f.write(f"- **CSV数据文件**: `china_food_simplified.csv`\n")
        f.write(f"- **本报告**: `数据处理报告_v2.md`\n")

    print(f"\n报告已生成: {report_path}")
    return report_path


def main():
    all_foods, stats = extract_data()

    csv_path = generate_csv(all_foods, stats)

    excel_path = generate_excel_with_summary(all_foods, stats)

    report_path = generate_review_report(all_foods, stats)

    print("\n" + "=" * 60)
    print("数据处理完成！")
    print("=" * 60)
    print(f"\n输出目录: {OUTPUT_DIR}")
    print(f"  - 审核用Excel: {os.path.basename(excel_path) if excel_path else '未生成'}")
    print(f"  - CSV数据文件: {os.path.basename(csv_path)}")
    print(f"  - 处理报告: {os.path.basename(report_path)}")


if __name__ == "__main__":
    main()