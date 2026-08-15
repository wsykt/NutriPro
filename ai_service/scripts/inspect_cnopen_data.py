"""检查CnOpenData中国食物营养成分数据Excel文件结构"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("正在安装openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

BASE_DIR = r"C:\Users\13425\Desktop\个人健康助手\CnOpenData中国食物营养成分数据"
FIELD_FILE = os.path.join(BASE_DIR, "CnOpenData中国食物营养成分数据-字段表.xlsx")
CONTENT_FILE = os.path.join(BASE_DIR, "CnOpenData中国食物营养成分数据（样本）", "营养成分含量信息表.xlsx")
SCORE_FILE = os.path.join(BASE_DIR, "CnOpenData中国食物营养成分数据（样本）", "营养成分得分信息表.xlsx")


def inspect_file(file_path, name, max_rows=10):
    print("=" * 70)
    print(f"文件: {name}")
    print(f"路径: {file_path}")
    print("=" * 70)

    if not os.path.exists(file_path):
        print(f"  [错误] 文件不存在!")
        return

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"工作表名称: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- 工作表: {sheet_name} ---")
        rows = list(ws.iter_rows(values_only=True))
        print(f"总行数: {len(rows)}")

        if not rows:
            continue

        print(f"\n表头(第1行), 共{len(rows[0])}列:")
        for i, header in enumerate(rows[0]):
            print(f"  列{i+1}: {header}")

        print(f"\n前{min(max_rows, len(rows)-1)}行数据示例:")
        for r_idx, row in enumerate(rows[1:max_rows+1], start=2):
            print(f"  行{r_idx}: {row}")

    wb.close()


if __name__ == "__main__":
    inspect_file(FIELD_FILE, "字段表")
    print("\n\n")
    inspect_file(CONTENT_FILE, "营养成分含量信息表")
    print("\n\n")
    inspect_file(SCORE_FILE, "营养成分得分信息表")
